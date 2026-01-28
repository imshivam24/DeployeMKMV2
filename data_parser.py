"""
Optimized data extraction that reads Excel once and caches all data.
Formulas are stored and re-evaluated with new pH/V values.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math
import copy
import re

logger = logging.getLogger(__name__)


class CachedExcelDataProcessor:
    """
    Optimized Excel processor that reads data once and caches it.
    Stores formulas and re-evaluates them with new pH/V values.
    """

    def __init__(self, excel_path: str):
        """Initialize and load all data from Excel file once."""
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        # Cache all data AND formulas on initialization
        self._cached_data = self._load_all_data()
        self._cached_formulas = self._load_formulas()
        logger.info(f"Excel data and formulas cached from {excel_path}")

    def _evaluate_formula(self, formula_data: Any, pH: float, V: float, row: int = None) -> float:
        """
        Evaluate a formula string with given pH and V values.
        Handles both 'Local Environment' references and internal cell references.
        
        Args:
            formula_data: Dict with 'formula' or 'value', or direct value
            pH: pH value to substitute
            V: Potential value to substitute
            row: Row number for looking up relative cell references
            
        Returns:
            Evaluated numeric result
        """
        # Handle dictionary format from new loader
        if isinstance(formula_data, dict):
            if 'value' in formula_data:
                try:
                    return float(formula_data['value'])
                except (ValueError, TypeError):
                    return 0.0
            formula = formula_data.get('formula')
            row = formula_data.get('row', row)
        else:
            formula = formula_data
        
        # If not a formula, return as-is
        if not isinstance(formula, str) or not formula.startswith('='):
            try:
                return float(formula) if formula is not None else 0.0
            except (ValueError, TypeError):
                return 0.0
        
        # Remove the '=' prefix
        formula_str = formula[1:]
        
        # Step 1: Replace 'Local Environment' sheet references with pH/V values
        env_headers = self._cached_formulas.get('LocalEnv_headers', {})
        
        # Replace V references
        if 'V' in env_headers:
            v_col = env_headers['V']
            formula_str = re.sub(
                r"'Local Environment'!\$?" + v_col + r"\$?\d+",
                str(V),
                formula_str
            )
        
        # Replace pH references
        if 'pH' in env_headers:
            ph_col = env_headers['pH']
            formula_str = re.sub(
                r"'Local Environment'!\$?" + ph_col + r"\$?\d+",
                str(pH),
                formula_str
            )
        
        # Step 2: Replace internal cell references (like C2, E2) with cached values
        # Get cached cell values from Reactions sheet
        reactions_cells = self._cached_formulas.get('Reactions_cells', {})
        
        # Find all cell references (e.g., C2, E3, AA10)
        cell_pattern = r'\b([A-Z]+)(\d+)\b'
        
        def replace_cell_ref(match):
            col_letter = match.group(1)
            cell_row = match.group(2)
            cell_ref = f"{col_letter}{cell_row}"
            
            # Look up the value in cached cells
            if cell_ref in reactions_cells:
                value = reactions_cells[cell_ref]
                # If the referenced cell also has a formula that depends on V/pH,
                # we need to handle it recursively, but for now use the cached value
                return str(value)
            else:
                logger.warning(f"Cell reference {cell_ref} not found in cache")
                return "0"
        
        formula_str = re.sub(cell_pattern, replace_cell_ref, formula_str)
        
        if len(formula[:50]) < 80:
            logger.debug(f"Original: {formula[:80]}")
        logger.debug(f"After sub: {formula_str[:100]}")
        
        # Create a safe evaluation context
        context = {
            'abs': abs,
            'exp': math.exp,
            'log': math.log,
            'log10': math.log10,
            'sqrt': math.sqrt,
            'pow': pow,
            'min': min,
            'max': max,
        }
        
        try:
            # Evaluate the expression
            result = eval(formula_str, {"__builtins__": {}}, context)
            return float(result)
            
        except Exception as e:
            logger.warning(f"Failed to evaluate formula '{formula[:80]}...': {e}")
            logger.debug(f"After substitution: '{formula_str[:100]}'")
            return 0.0
        """
        Load all sheets and evaluate formulas once.
        Returns dictionary of DataFrames for each sheet.
        """
        try:
            # Read all sheets with data_only=False to preserve formulas
            excel_file = pd.ExcelFile(self.excel_path, engine='openpyxl')
            
            cached = {}
            for sheet_name in excel_file.sheet_names:
                # Read with formulas evaluated
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                cached[sheet_name] = df.copy()
                logger.debug(f"Cached sheet '{sheet_name}': {df.shape}")
            
            return cached
            
        except Exception as e:
            logger.error(f"Error loading Excel data: {e}")
            raise

    def _load_formulas(self) -> Dict[str, Dict[str, List]]:
        """
        Load formula strings AND all cell values from Excel.
        Returns dict with formula strings and cell values for reference.
        """
        try:
            wb = load_workbook(self.excel_path, data_only=False)
            wb_values = load_workbook(self.excel_path, data_only=True)  # Load with evaluated values
            
            formulas = {}
            
            # Store the Local Environment sheet structure for reference lookups
            if 'Local Environment' in wb.sheetnames:
                ws_env = wb['Local Environment']
                formulas['LocalEnv_headers'] = {}
                formulas['LocalEnv_row'] = 2  # Data is typically in row 2
                
                # Map column names to column letters
                for col_idx, cell in enumerate(ws_env[1], start=1):
                    if cell.value:
                        formulas['LocalEnv_headers'][cell.value] = get_column_letter(col_idx)
                
                logger.debug(f"Local Environment columns: {formulas['LocalEnv_headers']}")
            
            # Load formulas AND cell values from Reactions sheet
            if 'Reactions' in wb.sheetnames:
                ws = wb['Reactions']
                ws_values = wb_values['Reactions']
                
                # Store all cell values by column letter and row for reference
                formulas['Reactions_cells'] = {}
                for row_idx in range(1, ws.max_row + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(col_idx)
                        cell_ref = f"{col_letter}{row_idx}"
                        cell_value = ws_values.cell(row=row_idx, column=col_idx).value
                        if cell_value is not None:
                            formulas['Reactions_cells'][cell_ref] = cell_value
                
                logger.debug(f"Cached {len(formulas['Reactions_cells'])} cell values from Reactions sheet")
                
                # Find column indices for G_f, G_b, DelG_rxn
                header_row = 1
                col_map = {}
                for col_idx, cell in enumerate(ws[header_row], start=1):
                    if cell.value in ['G_f', 'G_b', 'DelG_rxn']:
                        col_map[cell.value] = col_idx
                
                # Extract formulas for each column
                formulas['Reactions'] = {
                    'G_f': [],
                    'G_b': [],
                    'DelG_rxn': [],
                    'G_f_col': col_map.get('G_f'),
                    'G_b_col': col_map.get('G_b'),
                    'DelG_rxn_col': col_map.get('DelG_rxn'),
                }
                
                max_row = ws.max_row
                for row_idx in range(2, max_row + 1):  # Start from row 2 (after header)
                    for col_name, col_idx in col_map.items():
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            # Store formula string if it exists, otherwise store the value
                            if isinstance(cell.value, str) and cell.value.startswith('='):
                                formulas['Reactions'][col_name].append({
                                    'formula': cell.value,
                                    'row': row_idx
                                })
                                if row_idx <= 3:  # Log first few
                                    logger.debug(f"Row {row_idx}, {col_name}: {cell.value[:80]}")
                            else:
                                # Not a formula, store the value
                                formulas['Reactions'][col_name].append({
                                    'value': cell.value,
                                    'row': row_idx
                                })
                        else:
                            formulas['Reactions'][col_name].append(None)
                
                logger.info(f"Loaded {len(formulas['Reactions']['G_f'])} barrier formulas from Reactions sheet")
            
            wb.close()
            wb_values.close()
            return formulas
            
        except Exception as e:
            logger.error(f"Error loading formulas: {e}")
            import traceback
            traceback.print_exc()
            # Return empty dict if formula loading fails
            return {}

    def _load_all_data(self) -> Dict[str, pd.DataFrame]:
        """
        Load all sheets and evaluate formulas once.
        Returns dictionary of DataFrames for each sheet.
        """
        try:
            # Read all sheets with data_only=True to get formula results
            excel_file = pd.ExcelFile(self.excel_path, engine='openpyxl')
            
            cached = {}
            for sheet_name in excel_file.sheet_names:
                # Read with formulas evaluated
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                cached[sheet_name] = df.copy()
                logger.debug(f"Cached sheet '{sheet_name}': {df.shape}")
            
            return cached
            
        except Exception as e:
            logger.error(f"Error loading Excel data: {e}")
            raise

    def get_data_for_conditions(self, pH: float, V: float) -> Dict[str, Any]:
        """
        Get data for specific pH and V conditions.
        Works entirely with cached data - no Excel file access.
        
        Args:
            pH: pH value
            V: Potential value
            
        Returns:
            Dictionary containing all extracted data
        """
        try:
            # Work with copies to avoid modifying cached data
            df_reactions = self._cached_data['Reactions'].copy()
            df_local_env = self._cached_data['Local Environment'].copy()
            df_species = self._cached_data['Input-Output Species'].copy()
            
            # CRITICAL: Update pH and V for all rows in local environment
            df_local_env['pH'] = pH
            df_local_env['V'] = V
            
            logger.debug(f"Setting conditions: pH={pH}, V={V}")
            
            # Extract reaction data
            reactions = df_reactions['Reactions'].tolist()
            
            # Check if we have formulas cached, if so, evaluate them with pH/V
            if self._cached_formulas and 'Reactions' in self._cached_formulas:
                logger.debug(f"Re-evaluating formulas with pH={pH}, V={V}")
                
                # First, evaluate G_f and G_b (which depend on V directly)
                # and update the cell cache so DelG_rxn can reference them
                Ea_raw = []
                g_f_col = get_column_letter(self._cached_formulas['Reactions'].get('G_f_col', 3))
                for i, formula_data in enumerate(self._cached_formulas['Reactions']['G_f']):
                    row_num = i + 2  # Rows start at 2 (after header)
                    ea_value = self._evaluate_formula(formula_data, pH, V)
                    Ea_raw.append(ea_value)
                    # Update cache so DelG_rxn formulas can reference this
                    self._cached_formulas['Reactions_cells'][f"{g_f_col}{row_num}"] = ea_value
                
                Eb_raw = []
                g_b_col = get_column_letter(self._cached_formulas['Reactions'].get('G_b_col', 4))
                for i, formula_data in enumerate(self._cached_formulas['Reactions']['G_b']):
                    row_num = i + 2
                    eb_value = self._evaluate_formula(formula_data, pH, V)
                    Eb_raw.append(eb_value)
                    # Update cache
                    self._cached_formulas['Reactions_cells'][f"{g_b_col}{row_num}"] = eb_value
                
                # Now evaluate DelG_rxn (which may reference C2, D2, etc.)
                delE_list = []
                for i, formula_data in enumerate(self._cached_formulas['Reactions']['DelG_rxn']):
                    delE_list.append(self._evaluate_formula(formula_data, pH, V))
                
                logger.debug(f"Sample raw values before assignment:")
                logger.debug(f"  Ea[0:3] = {Ea_raw[:3]}")
                logger.debug(f"  Eb[0:3] = {Eb_raw[:3]}")
                logger.debug(f"  DelG[0:3] = {delE_list[:3]}")
            else:
                # No formulas cached, use values from DataFrame
                logger.warning("No formulas cached - using static values from Excel")
                Ea_raw = []
                for val in df_reactions['G_f'].tolist():
                    try:
                        Ea_raw.append(float(val) if val is not None else 0.0)
                    except (ValueError, TypeError):
                        Ea_raw.append(0.0)
                
                Eb_raw = []
                for val in df_reactions['G_b'].tolist():
                    try:
                        Eb_raw.append(float(val) if val is not None else 0.0)
                    except (ValueError, TypeError):
                        Eb_raw.append(0.0)
                
                delE_list = []
                for val in df_reactions['DelG_rxn'].tolist():
                    try:
                        delE_list.append(float(val) if val is not None else 0.0)
                    except (ValueError, TypeError):
                        delE_list.append(0.0)

                        
            for i in range(len(Ea_raw)):
                try:
                    ea = float(Ea_raw[i])
                    dg = float(delE_list[i])
                except (TypeError, ValueError, IndexError):
                    continue

                # Only overwrite Gb if DelG is valid
                Eb_raw[i] = ea - dg            
            
            # Apply barrier assignment logic
            Ea, Eb = self._assign_barriers_with_safety(Ea_raw, Eb_raw, delE_list)
            
            # Validate no negative barriers
            negative_count = sum(1 for i in range(len(Ea)) if Ea[i] < 0 or Eb[i] < 0)
            if negative_count > 0:
                logger.error(f"❌ {negative_count} negative barriers still present!")
            else:
                logger.debug(f"✅ All {len(Ea)} activation barriers are non-negative")
            
            # Extract environment data - USE THE PARAMETERS, NOT FROM DATAFRAME
            # We pass pH and V as arguments, so use those directly
            V_val = float(V)
            pH_val = float(pH)
            P = float(df_local_env['Pressure'].iloc[0])
            
            logger.debug(f"Using conditions: pH={pH_val}, V={V_val}, P={P}")
            
            # Extract species data
            gases = df_species['Species'].tolist()
            
            # Convert concentrations to float
            concentrations = []
            for val in df_species['Input MKMCXX'].tolist():
                try:
                    concentrations.append(float(val) if val is not None else 0.0)
                except (ValueError, TypeError):
                    concentrations.append(0.0)
            
            # Parse reactions
            parsed_reactions = self._parse_reactions(reactions)
            
            # Find adsorbates
            adsorbates = self._extract_adsorbates(parsed_reactions)
            
            return {
                'reactions': reactions,
                'Ea': Ea,
                'Eb': Eb,
                'V': V_val,  # Use the parameter value
                'pH': pH_val,  # Use the parameter value
                'P': P,
                'gases': gases,
                'concentrations': concentrations,
                'adsorbates': adsorbates,
                'activity': np.zeros(len(adsorbates)),
                **parsed_reactions
            }
            
        except Exception as e:
            logger.error(f"Error extracting data for pH={pH}, V={V}: {e}")
            raise

    def _assign_barriers_with_safety(self, Ea: List[float], Eb: List[float], 
                                    delE_list: List[float]) -> Tuple[List[float], List[float]]:
        """
        Apply assignment logic with safety clamping to prevent negative barriers.
        """
        Ea_final = Ea.copy()
        Eb_final = Eb.copy()
        delE_index = 0

        for i in range(len(Ea_final)):
            # Handle None/NaN and convert to float
            try:
                ea_val = float(Ea_final[i]) if Ea_final[i] is not None else 0.0
                if math.isnan(ea_val):
                    ea_val = 0.0
            except (ValueError, TypeError):
                ea_val = 0.0
            
            try:
                eb_val = float(Eb_final[i]) if Eb_final[i] is not None else 0.0
                if math.isnan(eb_val):
                    eb_val = 0.0
            except (ValueError, TypeError):
                eb_val = 0.0

            # Get DelG_rxn value
            delE = None
            if delE_index < len(delE_list):
                try:
                    delE = float(delE_list[delE_index])
                    if math.isnan(delE):
                        delE = None
                    else:
                        delE_index += 1
                except (ValueError, TypeError):
                    delE = None

            # Apply assignment logic
            if ea_val == 0 and eb_val == 0:
                if delE is not None:
                    if delE > 0:
                        Ea_final[i] = delE
                        Eb_final[i] = 0.0
                    else:
                        Ea_final[i] = 0.0
                        Eb_final[i] = -delE
            elif ea_val < 0:
                if delE is not None:
                    Ea_final[i] = 0.0
                    Eb_final[i] = -delE
                else:
                    Ea_final[i] = 0.0
            elif eb_val < 0:
                if delE is not None:
                    Ea_final[i] = delE
                    Eb_final[i] = 0.0
                else:
                    Eb_final[i] = 0.0
            
            # Safety clamping - ensure float type
            Ea_final[i] = float(max(0.0, Ea_final[i]))
            Eb_final[i] = float(max(0.0, Eb_final[i]))

        return Ea_final, Eb_final

    def _parse_reactions(self, reactions: List[str]) -> Dict[str, List[str]]:
        """Parse reaction strings into reactants and products."""
        reactant1, reactant2, reactant3 = [], [], []
        product1, product2, product3 = [], [], []

        for rxn in reactions:
            try:
                # Handle both arrow types
                if "→" in rxn:
                    reactants_str, products_str = rxn.strip().split("→")
                else:
                    reactants_str, products_str = rxn.strip().split("->")

                # Parse reactants
                reactants = [r.strip() for r in reactants_str.split("+")]
                reactant1.append(f"{{{reactants[0]}}}")
                reactant2.append(f"{{{reactants[1]}}}" if len(reactants) > 1 else "")
                reactant3.append(f"{{{reactants[2]}}}" if len(reactants) > 2 else "")

                # Parse products
                products = [p.strip() for p in products_str.split("+")]
                product1.append(f"{{{products[0]}}}")
                product2.append(f"{{{products[1]}}}" if len(products) > 1 else "")
                product3.append(f"{{{products[2]}}}" if len(products) > 2 else "")

            except Exception as e:
                logger.error(f"Error parsing reaction '{rxn}': {e}")
                for lst in [reactant1, reactant2, reactant3, product1, product2, product3]:
                    lst.append("")

        return {
            'Reactant1': reactant1,
            'Reactant2': reactant2,
            'Reactant3': reactant3,
            'Product1': product1,
            'Product2': product2,
            'Product3': product3
        }

    def _extract_adsorbates(self, parsed_reactions: Dict[str, List[str]]) -> List[str]:
        """Extract unique adsorbates from parsed reactions."""
        adsorbates = set()

        for key in ['Reactant1', 'Reactant2', 'Product1', 'Product2']:
            for item in parsed_reactions.get(key, []):
                if "*" in item and item != "":
                    species = item.strip("{}").strip()
                    if species != "*":
                        adsorbates.add(species)

        return list(adsorbates)


def data_extract(pH: float, V: float, processor: CachedExcelDataProcessor) -> Tuple:
    """
    Optimized data extraction using cached processor.
    
    Args:
        pH: pH value
        V: Potential value
        processor: Pre-initialized CachedExcelDataProcessor instance
    
    Returns:
        Tuple containing all extracted data
    """
    # Extract data from cached processor (no file I/O)
    data = processor.get_data_for_conditions(pH, V)
    
    return (
        data['gases'],
        data['concentrations'],
        data['adsorbates'],
        data['activity'],
        data['Reactant1'],
        data['Reactant2'],
        data['Reactant3'],
        data['Product1'],
        data['Product2'],
        data['Product3'],
        data['Ea'],
        data['Eb'],
        data['P'],
        data['reactions']
    )